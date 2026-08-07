"""
Spreadsheet parsing service — turns an uploaded .xlsx or .csv file into
structured tables (headers + rows), and produces a capped text
rendering of each table for embedding/retrieval.

Deliberately NOT trying to be a general spreadsheet engine: no formula
evaluation (values are read as-computed by whatever last saved the
file), no merged-cell reconstruction beyond taking the anchor cell's
value, no multi-header-row detection. The first non-empty row of each
sheet is treated as the header row. Good enough for the flat,
single-header-row tables NGO budget/beneficiary/monitoring sheets
typically are; a more sophisticated parser is a real v2 problem, not
something to half-solve here.
"""

import csv
import io
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal

MAX_ROWS_PER_SHEET = 500  # parsed & stored in full up to this cap
MAX_ROWS_IN_CHUNK_PREVIEW = 50  # shown in the retrieval-chunk markdown rendering


@dataclass
class ParsedTable:
    sheet_name: str
    headers: list[str]
    rows: list[list]  # each inner list aligned to `headers` by position
    truncated: bool  # True if the sheet had more than MAX_ROWS_PER_SHEET rows


def _json_safe(value):
    """Coerces a cell value into something JSONB can store directly."""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def _is_blank_row(row: list) -> bool:
    return all(cell is None or str(cell).strip() == "" for cell in row)


def parse_xlsx(file_bytes: bytes) -> list[ParsedTable]:
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    tables: list[ParsedTable] = []

    for worksheet in workbook.worksheets:
        row_iter = worksheet.iter_rows(values_only=True)

        headers: list[str] | None = None
        data_rows: list[list] = []
        truncated = False

        for raw_row in row_iter:
            row = list(raw_row)
            if _is_blank_row(row):
                continue

            if headers is None:
                # First non-empty row is the header row. Blank header
                # cells get a positional fallback name so every column
                # still has something usable to reference.
                headers = [
                    str(cell).strip() if cell not in (None, "") else f"Column {i + 1}"
                    for i, cell in enumerate(row)
                ]
                continue

            if len(data_rows) >= MAX_ROWS_PER_SHEET:
                truncated = True
                break

            # Pad/truncate the row to match the header count — real
            # spreadsheets are rarely perfectly rectangular.
            safe_row = [_json_safe(c) for c in row[: len(headers)]]
            safe_row += [None] * (len(headers) - len(safe_row))
            data_rows.append(safe_row)

        if headers is not None and data_rows:
            tables.append(
                ParsedTable(
                    sheet_name=worksheet.title,
                    headers=headers,
                    rows=data_rows,
                    truncated=truncated,
                )
            )

    return tables


def parse_csv(file_bytes: bytes, fallback_name: str) -> list[ParsedTable]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))

    rows = [row for row in reader if not _is_blank_row(row)]
    if not rows:
        return []

    headers = [cell.strip() if cell else f"Column {i + 1}" for i, cell in enumerate(rows[0])]
    truncated = len(rows) - 1 > MAX_ROWS_PER_SHEET
    data_rows = rows[1 : 1 + MAX_ROWS_PER_SHEET]

    safe_rows = []
    for row in data_rows:
        safe_row = list(row[: len(headers)])
        safe_row += [None] * (len(headers) - len(safe_row))
        safe_rows.append(safe_row)

    return [
        ParsedTable(sheet_name=fallback_name, headers=headers, rows=safe_rows, truncated=truncated)
    ]


def parse_spreadsheet(file_bytes: bytes, doc_type: str, filename: str) -> list[ParsedTable]:
    if doc_type == "xlsx":
        return parse_xlsx(file_bytes)
    if doc_type == "csv":
        return parse_csv(file_bytes, fallback_name=filename)
    raise ValueError(f"Unsupported spreadsheet type: {doc_type}")


def render_table_markdown(table: ParsedTable) -> str:
    """
    Capped markdown rendering used as the retrieval CHUNK's content —
    this is what gets embedded and what the LLM actually reads, so it
    stays small regardless of how large the full stored table is.
    The full data always remains available in SpreadsheetTable.rows
    for the comparison tool, unaffected by this cap.
    """
    preview_rows = table.rows[:MAX_ROWS_IN_CHUNK_PREVIEW]

    header_line = "| " + " | ".join(str(h) for h in table.headers) + " |"
    separator_line = "| " + " | ".join("---" for _ in table.headers) + " |"
    body_lines = [
        "| " + " | ".join("" if c is None else str(c) for c in row) + " |" for row in preview_rows
    ]

    note = ""
    if len(table.rows) > MAX_ROWS_IN_CHUNK_PREVIEW:
        note = f"\n\n_Showing {MAX_ROWS_IN_CHUNK_PREVIEW} of {len(table.rows)} rows._"

    return f"Sheet: {table.sheet_name}\n\n" + "\n".join([header_line, separator_line, *body_lines]) + note